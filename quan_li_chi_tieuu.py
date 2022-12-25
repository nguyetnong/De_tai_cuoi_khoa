from tkinter import*
from tkinter import ttk
from tkcalendar import Calendar,DateEntry

from datetime import date
import openpyxl

list_chitieu=[]
def Read_Excel():
    sum=0.0
    wb=openpyxl.load_workbook(r"D:\Nguyet\bai_cuoi_khoa\QUAN_LI_CHI_TIEU.xlsx")
    ws=wb["quản lí chi tiêu"]
    count_row=ws.max_row
    tv.delete(*tv.get_children())
    list_chitieu.clear()
    for row in range(2,count_row+1):
        cell_ID = ws.cell(row=row,column=1).value
        cell_tenchiTieu = ws.cell(row=row,column=2).value
        cell_sotien = ws.cell(row=row,column=3).value
        cell_thogian = ws.cell(row=row,column=4).value
        cell_trangthai = ws.cell(row=row,column=5).value
        if cell_ID != None:
            khoan_chi={"id":cell_ID,"tenchitieu":cell_tenchiTieu,"sotien":cell_sotien,"thoigian":cell_thogian ,"trangthai":cell_trangthai}
            list_chitieu.append(khoan_chi)
            if cell_trangthai != "D":
                tv.insert(parent='', index=cell_ID, iid=cell_ID, text='', values=(cell_ID,cell_tenchiTieu,"{:,}".format(float(cell_sotien)),cell_thogian))
                sum=sum+float(cell_sotien)
    label_value['text'] = "{:,}".format(sum)

    wb.close()

def value_format(num):
    return "{:,0f}".format(num)

def save_excell():
    value_format
    wb=openpyxl.load_workbook(r"D:\Nguyet\bai_cuoi_khoa\QUAN_LI_CHI_TIEU.xlsx")
    ws=wb["quản lí chi tiêu"]
    count_row=ws.max_row
    for row in range(2,count_row+1):
        ws.cell(row=row,column=1).value = ""
        ws.cell(row=row,column=2).value = ""
        ws.cell(row=row,column=3).value = ""
        ws.cell(row=row,column=4).value = ""
        ws.cell(row=row,column=5).value = ""

    row =2
    for item in list_chitieu:
        ID = item["id"]
        tenchitieu = item["tenchitieu"]
        sotien =item["sotien"]
        thoigian = item['thoigian']
        trangthai = item['trangthai']
        ws.cell(row=row,column=1).value = ID
        ws.cell(row=row,column=2).value = tenchitieu
        ws.cell(row=row,column=3).value = sotien
        ws.cell(row=row,column=4).value = thoigian
        ws.cell(row=row,column=5).value = trangthai
        row +=1
    wb.save(r"D:\Nguyet\bai_cuoi_khoa\QUAN_LI_CHI_TIEU.xlsx")
    wb.close()

def them_chitieu():
    max=0
    
    for item in list_chitieu:
        if int(item["id"]) > max:
            max = int(item["id"])

    ID = max + 1
    tenchitieu=entry_tenchitieu.get()
    sotien=entry_sotien.get()
    thoigian=entry_thoigian.get()
    khoan_chi={"id": ID,"tenchitieu":tenchitieu,"sotien":sotien,"thoigian":thoigian,"trangthai":""}
    list_chitieu.append(khoan_chi)

    tv.insert(parent='', index=ID, iid=ID, text='', values=(ID,tenchitieu,"{:,}".format(float(sotien)),thoigian))

    save_excell()
    Read_Excel()
    clearForm()
    UnLock_Widget()

def clearForm():
    UnLock_Widget()
    entry_tenchitieu.delete(0,END)
    entry_sotien.delete(0,END)
    entry_thoigian.delete(0,END)
    entry_id.delete(0,END)
    Lock_Widget()
    
def Lock_Widget():
    entry_id.configure(state='readonly')
    entry_tenchitieu.configure(state='readonly')
    entry_thoigian.configure(state='readonly')
    entry_sotien.configure(state='readonly')

def UnLock_Widget():
    entry_id.configure(state='normal')
    entry_tenchitieu.configure(state='normal')
    entry_thoigian.configure(state='normal')
    entry_sotien.configure(state='normal')


def xem():
    tv.delete(*tv.get_children())
    sum=0.0
    if combobox_Mode_xem.get() =="xem theo id":
        item=int(entry_id.get())
        for khoan_chi in list_chitieu:
            if item==khoan_chi["id"] and khoan_chi["trangthai"] != "D":
                tv.insert(parent='', index=item, iid=item, text='', values=(item,khoan_chi["tenchitieu"],"{:,}".format(float(khoan_chi["sotien"])),khoan_chi["thoigian"]))
                sum=sum+float(khoan_chi["sotien"])

    elif combobox_Mode_xem.get() =="xem theo tên":
        item=entry_tenchitieu.get()
        for khoan_chi in list_chitieu:
            if item==khoan_chi["tenchitieu"] and khoan_chi["trangthai"] != "D":
                tv.insert(parent='', index=int(khoan_chi["id"]), iid=int(khoan_chi["id"]), text='', values=(khoan_chi["id"],item,"{:,}".format(float(khoan_chi["sotien"])),khoan_chi["thoigian"]))
                sum=sum+float(khoan_chi["sotien"])
                
    elif combobox_Mode_xem.get() =="xem theo ngày":
        item=entry_thoigian.get()
        for khoan_chi in list_chitieu:
            if item==khoan_chi["thoigian"] and khoan_chi["trangthai"] != "D":
                tv.insert(parent='', index=int(khoan_chi["id"]), iid=int(khoan_chi["id"]), text='', values=(khoan_chi["id"],khoan_chi["tenchitieu"],"{:,}".format(float(khoan_chi["sotien"])),item))
                sum=sum+float(khoan_chi["sotien"])
    
    elif combobox_Mode_xem.get() =="xem tất cả":
        for khoan_chi in list_chitieu:
            if khoan_chi["trangthai"] != "D":
                tv.insert(parent='', index=int(khoan_chi["id"]), iid=int(khoan_chi["id"]), text='', values=(khoan_chi["id"],khoan_chi["tenchitieu"],"{:,}".format(float(khoan_chi["sotien"])),khoan_chi["thoigian"]))  
                sum=sum+float(khoan_chi["sotien"])
    label_value['text'] = "{:,}".format(sum)

def update():
    id_update=int(entry_id.get())
    index=0
    for khoan_chi in list_chitieu:    
        if id_update == khoan_chi["id"] and khoan_chi["trangthai"] != "D":           
            tenchitieu_new=entry_tenchitieu.get()
            list_chitieu[index]["tenchitieu"]=tenchitieu_new         
           
            sotien_new=entry_sotien.get()
            list_chitieu[index]["sotien"]=sotien_new
            
            thoigian_new=entry_thoigian.get()
            list_chitieu[index]["thoigian"]=thoigian_new

            khoan_chi={"tenchitieu":tenchitieu_new,"sotien":sotien_new,"thoigian":thoigian_new,"trangthai":""}
            selected = tv.focus()
            tv.item(selected, values=(id_update,tenchitieu_new,"{:,}".format(float(sotien_new)),thoigian_new))
        else:
            index+=1
    save_excell()
    clearForm()
    UnLock_Widget()

def xoa():
    if combobox_Mode_Xoa.get() =="xóa tất cả":
        for i in range(len(list_chitieu)):
            list_chitieu[i]["trangthai"] ="D"
            
    elif combobox_Mode_Xoa.get() =="xóa theo id":
        id_canxoa=int(entry_id.get())
        idx = 0
        for khoan_chi in list_chitieu:
            if id_canxoa==khoan_chi["id"]:
                list_chitieu[idx]["trangthai"] ="D"
            idx += 1
    elif combobox_Mode_Xoa.get() =="xóa theo tên":
        ten_canxoa=entry_tenchitieu.get()
        idx = 0
        for khoan_chi in list_chitieu:
            if ten_canxoa==khoan_chi["tenchitieu"]:
                list_chitieu[idx]["trangthai"] ="D"
            idx += 1
                
    elif combobox_Mode_Xoa.get() =="xóa theo ngày":
        id_canxoa= entry_thoigian.get()
        idx = 0
        for khoan_chi in list_chitieu:
            if id_canxoa==khoan_chi["thoigian"]:
                list_chitieu[idx]["trangthai"] ="D"
            idx += 1
    save_excell()
    Read_Excel()

def OnDoubleClick(event):
    clearForm()
    selected = tv.focus()
    temp = tv.item(selected, 'values')
    ID =int(temp[0])
   
    UnLock_Widget()
    for khoan_chi in list_chitieu:
       
        if khoan_chi["id"]==ID:
            tenchitieu = khoan_chi["tenchitieu"]
            sotien = khoan_chi["sotien"]
            ngaychitieu = khoan_chi["thoigian"]
            
            entry_tenchitieu.insert(END,tenchitieu)
            entry_sotien.insert(END,sotien)
            entry_thoigian.insert(END,ngaychitieu)
            entry_id.insert(END,ID)
            entry_id.configure(state='readonly')
            break
def ComboboxSelected_Xoa(eventObject):
    clearForm()
    if combobox_Mode_Xoa.get() =="xóa theo id":
        entry_id.configure(state='normal')
        entry_tenchitieu.configure(state='readonly')
        entry_thoigian.configure(state='readonly')
        
    elif combobox_Mode_Xoa.get() =="xóa theo tên":
        entry_id.configure(state='readonly')
        entry_tenchitieu.configure(state='normal')
        entry_thoigian.configure(state='readonly')
        
    elif combobox_Mode_Xoa.get() =="xóa theo ngày":
        entry_id.configure(state='readonly')
        entry_tenchitieu.configure(state='readonly')
        entry_thoigian.configure(state='normal')
        
    elif combobox_Mode_Xoa.get() =="xóa tất cả":
        entry_id.configure(state='readonly')
        entry_tenchitieu.configure(state='readonly')
        entry_thoigian.configure(state='readonly')

def ComboboxSelected_Xem(eventObject):
    clearForm()
    if combobox_Mode_xem.get() =="xem theo id":
        entry_id.configure(state='normal')
        entry_tenchitieu.configure(state='readonly')
        entry_thoigian.configure(state='readonly')
        
    elif combobox_Mode_xem.get() =="xem theo tên":
        entry_id.configure(state='readonly')
        entry_tenchitieu.configure(state='normal')
        entry_thoigian.configure(state='readonly')
        
    elif combobox_Mode_xem.get() =="xem theo ngày":
        entry_id.configure(state='readonly')
        entry_tenchitieu.configure(state='readonly')
        entry_thoigian.configure(state='normal')
        
    elif combobox_Mode_xem.get() =="xem tất cả":
        entry_id.configure(state='readonly')
        entry_tenchitieu.configure(state='readonly')
        entry_thoigian.configure(state='readonly')

def  number(event):
    txt = event.char
    arr=['0','1','2','3','4','5','6','7','8','9','.'] #42424g
    value=entry_sotien.get()
    if txt not in arr:    
        entry_sotien.delete(len(value)-1,END)
    else:
        if txt=="." and value.count('.')>= 2:
            entry_sotien.delete(len(value)-1,END)
def new():
    clearForm()
    UnLock_Widget()


window=Tk()
window.eval('tk::PlaceWindow . center')
window.title("CHI TIÊU")
# Frame cho các Label-Entry input
Frame_Background_Input = Frame(master= window,highlightbackground="#474642", highlightthickness=1,padx=5,pady=5)
Frame_Background_Input.grid(row=0,column=0)

lable_tenchitieu=Label(master=Frame_Background_Input,text="Tên chi tiêu: ")
lable_tenchitieu.grid(row=0,column=0)
entry_tenchitieu=Entry(master=Frame_Background_Input)
entry_tenchitieu.grid(row=0,column=1)

label_id=Label(master=Frame_Background_Input,text="ID: ")
label_id.grid(row=0,column=2)
entry_id=Entry(master=Frame_Background_Input,width=5)
entry_id.grid(row=0,column=3)

label_sotien=Label(master=Frame_Background_Input,text="Số tiền: ")
label_sotien.grid(row=1,column=0)
entry_sotien=Entry(master=Frame_Background_Input)
entry_sotien.grid(row=1,column=1)
entry_sotien.bind("<KeyRelease>",number)

label_thoigian=Label(master=Frame_Background_Input,text="Ngày mua: ")
label_thoigian.grid(row=2,column=0)
entry_thoigian = DateEntry(Frame_Background_Input,width=17,bg="darkblue",fg="white",year=date.today().year,month=date.today().month,day=date.today().day,date_pattern='mm/dd/y')
entry_thoigian.grid(row=2,column=1)


# Frame cho các button chức năng
Frame_Background_Button = Frame(master= window,padx=5,pady=5)
Frame_Background_Button.grid(row=1,column=0)

Frame_Button_new =Frame(master=Frame_Background_Button,padx=5)
Frame_Button_new.grid(row=0,column=0)
Frame_Button_Them =Frame(master=Frame_Background_Button,padx=5)
Frame_Button_Them.grid(row=0,column=1)
Frame_Button_Sua =Frame(master=Frame_Background_Button,padx=5)
Frame_Button_Sua.grid(row=0,column=2)
Frame_Button_Xoa =Frame(master=Frame_Background_Button,padx=5)
Frame_Button_Xoa.grid(row=0,column=4)
Frame_Button_Xem =Frame(master=Frame_Background_Button,padx=5)
Frame_Button_Xem.grid(row=0,column=3)

Button_new=Button(master=Frame_Button_new,text="New",command=new,width=5 )
Button_new.grid(row=0,column=0)

Button_them=Button(master=Frame_Button_Them,text="Thêm",command=them_chitieu,width=5)
Button_them.grid(row=0,column=0)

Button_update=Button(master=Frame_Button_Sua,text="Sửa",command=update,width=5)
Button_update.grid(row=0,column=0)

combobox_Mode_Xoa = ttk.Combobox(Frame_Button_Xoa, width = 8)
combobox_Mode_Xoa['values'] = ('xóa tất cả','xóa theo id','xóa theo tên','xóa theo ngày')
combobox_Mode_Xoa.grid(row = 0, column = 0)
combobox_Mode_Xoa.current()
combobox_Mode_Xoa.bind("<<ComboboxSelected>>", ComboboxSelected_Xoa)
Button_xoa=Button(master=Frame_Button_Xoa,text="Xóa",command=xoa,width=5)
Button_xoa.grid(row=0,column=1)

combobox_Mode_xem = ttk.Combobox(Frame_Button_Xem, width = 8)
combobox_Mode_xem['values'] = ('xem theo id','xem theo tên','xem theo ngày','xem tất cả')
combobox_Mode_xem.grid(row=0,column=0)
combobox_Mode_xem.current()
combobox_Mode_xem.bind("<<ComboboxSelected>>", ComboboxSelected_Xem)
Button_xem=Button(master=Frame_Button_Xem,text="Xem ",command=xem,width=5)
Button_xem.grid(row=0,column=1)

# Frame background cho lưới (table)
style=ttk.Style()
style.configure("Trstyle.Treeview.Heading",bg="#ffedd4",foreground="red", font=("Timenewromans",10,"bold"))
Frame_Background_Grid = Frame(master=window)
Frame_Background_Grid.grid(row=3,column=0)
tv = ttk.Treeview(Frame_Background_Grid,style="Trstyle.Treeview")
tv['columns']=('ID', 'tenchitieu', 'sotien','ngaychitieu')
tv.column('#0', width=0, stretch=NO)
tv.column('ID', anchor=CENTER, width=40)
tv.column('tenchitieu', anchor=CENTER, width=120)
tv.column('sotien', anchor=CENTER, width=120)
tv.column('ngaychitieu', anchor=CENTER, width=120)

tv.heading('#0', text='', anchor=CENTER)
tv.heading('ID', text='ID', anchor=CENTER)
tv.heading('tenchitieu', text='Tên chi tiêu', anchor=CENTER)
tv.heading('sotien', text='Số tiền', anchor=CENTER)
tv.heading('ngaychitieu', text='Ngày chi tiêu', anchor=CENTER)

tv.grid(row=0,column=0)
tv.bind("<Double-1>", OnDoubleClick)

Frame_total=Frame(master=window)
Frame_total.grid(row=4,column=0)
label_total=Label(master=Frame_total,foreground="green",text="Total : ")
label_total.grid(row=0,column=0)
label_value=Label(master=Frame_total,width=15)
label_value.grid(row=0,column=1)



Read_Excel()

window.mainloop()